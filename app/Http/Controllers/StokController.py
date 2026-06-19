from datetime import datetime
from app.Models.Database import db_manager
from app.Models.SnapshotManager import SnapshotManager
from app.Services.SalesMonitoringService import SalesMonitoringService
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import request, jsonify, session, render_template, send_file



class StokController:
    """
    Controller untuk Stok Monitoring.
    Uses SnapshotManager for instant local searches.
    """

    # ──────────── Views ────────────

    @staticmethod
    def dashboard_page():
        """HTML Page: Business Dashboard"""
        server_key = session.get('selected_server')
        if not server_key:
            return render_template('index.html')
        return render_template('dashboard.html', dashboard_type='bisnis')

    @staticmethod
    def dashboard_stok_page():
        """HTML Page: Stok Dashboard"""
        server_key = session.get('selected_server')
        if not server_key:
            return render_template('index.html')
        return render_template('dashboard.html', dashboard_type='stok')

    @staticmethod
    def index_page():
        """HTML Page: Server selection"""
        return render_template('index.html')

    @staticmethod
    def monitoring_page():
        """HTML Page: Monitoring dashboard"""
        server_key = session.get('selected_server')
        tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))

        if not server_key:
            return render_template('index.html')

        return render_template('monitoring.html', tanggal=tanggal)

    @staticmethod
    def histori_page():
        """HTML Page: Cek histori pergerakan barang"""
        server_key = session.get('selected_server')
        if not server_key:
            return render_template('index.html')
        return render_template('histori.html')

    @staticmethod
    def perhitungan_stok_page():
        """HTML Page: Perhitungan Stok via Transaksi"""
        server_key = session.get('selected_server')
        if not server_key:
            return render_template('index.html')
        return render_template('perhitungan_stok.html')

    @staticmethod
    def monitoring_transaksi_page():
        """HTML Page: Cek barang dengan/tanpa transaksi"""
        server_key = session.get('selected_server')
        if not server_key:
            return render_template('index.html')
        return render_template('monitoring_transaksi.html')

    @staticmethod
    def monitoring_penjualan_page():
        """HTML Page: Monitoring Penjualan (On Demand)"""
        server_key = session.get('selected_server')
        if not server_key:
            return render_template('index.html')
        return render_template('monitoring_penjualan.html')

    @staticmethod
    def opname_page():
        """HTML Page: Opname Stok (Read-Only)"""
        server_key = session.get('selected_server')
        tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
        if not server_key:
            return render_template('index.html')
        return render_template('opname.html', tanggal=tanggal)

    @staticmethod
    def update_barang_page():
        """HTML Page: Update Barang (Harga & Status)"""
        server_key = session.get('selected_server')
        if not server_key:
            return render_template('index.html')
        servers = db_manager.get_available_servers()
        server = next((s for s in servers if s['key'] == server_key), None)
        server_type = server.get('type', 'grosir') if server else 'grosir'
        return render_template('update_barang.html', server_type=server_type)

    @staticmethod
    def mass_refresh_page():
        """HTML Page: Mass refresh semua server"""
        return render_template('mass_refresh.html')

    # ──────────── Server Session APIs ────────────

    @staticmethod
    def get_server_list():
        """API: Dapatkan list server yang available"""
        try:
            servers = db_manager.get_available_servers()
            
            # Filter berdasarkan akses jika bukan super_admin
            if session.get('role') != 'super_admin':
                allowed_servers = session.get('servers', [])
                servers = [s for s in servers if s['key'] in allowed_servers]
                
            return jsonify({'status': 'success', 'servers': servers})
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_dashboard_summary():
        """Ambil data summary untuk dashboard bisnis/stok"""
        server_key = session.get('selected_server')
        if not server_key:
            return jsonify({'success': False, 'message': 'Pilih server dulu'}), 400
        
        tahun = request.args.get('tahun')
        if tahun:
            try:
                tahun = int(tahun)
            except ValueError:
                tahun = None
        mode = request.args.get('mode', 'single')
        
        from app.Services.DashboardService import DashboardService
        
        if mode == 'global':
            result = DashboardService.get_global_summary(tahun)
        else:
            if not server_key:
                return jsonify({'status': 'error', 'message': 'No server selected'}), 400
            result = DashboardService.get_summary(server_key, tahun)
            
        if "error" in result:
            return jsonify({'status': 'error', 'message': result['error']}), 500
            
        return jsonify({'status': 'success', 'data': result})

    @staticmethod
    def fetch_sales_monitoring():
        """API: Ambil data monitoring penjualan on-demand"""
        server_key = session.get('selected_server')
        if not server_key:
            return jsonify({'success': False, 'message': 'Pilih server dulu'}), 400
            
        start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-%d'))
        end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
        view_name = request.args.get('view_name', 'mon_t_penjualan_per_nota')
        
        try:
            # AUTO-SYNC THE REQUESTED DATE RANGE ON-THE-FLY
            from app.Services.Snapshot.SalesDuckDBRunner import SalesDuckDBRunner
            SalesDuckDBRunner.sync_date_range(server_key, start_date, end_date)
            
            svc = SalesMonitoringService()
            result = svc.get_sales_data(server_key, start_date, end_date, view_name)
            return jsonify({'success': True, 'columns': result['columns'], 'data': result['data']})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @staticmethod
    def select_server():
        """Set server yang dipilih ke session"""
        try:
            data = request.get_json()
            server_key = data.get('server_key')

            servers = db_manager.get_available_servers()
            server_keys = [s['key'] for s in servers]

            if server_key not in server_keys:
                return jsonify({'status': 'error', 'message': f'Server key "{server_key}" tidak valid'}), 400
                
            if session.get('role') != 'super_admin':
                if server_key not in session.get('servers', []):
                    return jsonify({'status': 'error', 'message': f'Akses ke server "{server_key}" ditolak'}), 403

            session['selected_server'] = server_key
            session.modified = True

            server_name = next((s['name'] for s in servers if s['key'] == server_key), server_key)
            return jsonify({
                'status': 'success',
                'message': f'Server "{server_name}" dipilih',
                'server_key': server_key,
                'server_name': server_name
            })

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def get_current_server():
        """Get server yang sedang dipilih di session"""
        server_key = session.get('selected_server')
        if not server_key:
            return jsonify({'status': 'not_selected', 'message': 'Belum ada server yang dipilih'})

        servers = db_manager.get_available_servers()
        server = next((s for s in servers if s['key'] == server_key), None)
        return jsonify({'status': 'success', 'server': server})

    # ──────────── Snapshot APIs ────────────

    @staticmethod
    def trigger_refresh():
        """API: Trigger snapshot refresh for current server"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
            result = SnapshotManager.trigger_refresh(server_key, tanggal)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def trigger_delta_refresh():
        """API: Quick update — only fetch new transactions since last refresh"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
            result = SnapshotManager.trigger_delta_refresh(server_key, tanggal)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def trigger_weekly_refresh():
        """API: Weekly update — fetch transactions from the last 7 days"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
            result = SnapshotManager.trigger_weekly_refresh(server_key, tanggal)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def trigger_yearly_refresh():
        """API: Yearly update"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
            result = SnapshotManager.trigger_yearly_refresh(server_key, tanggal)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def snapshot_status():
        """API: Get snapshot status for current server"""
        server_key = session.get('selected_server')
        if not server_key:
            return jsonify({'state': 'empty', 'has_snapshot': False})

        status = SnapshotManager.get_status(server_key)
        return jsonify(status)

    @staticmethod
    def cancel_refresh():
        """API: Cancel running snapshot refresh"""
        server_key = session.get('selected_server')
        if not server_key:
            return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

        result = SnapshotManager.cancel_refresh(server_key)
        return jsonify(result)

    @staticmethod
    def trigger_refresh_target(server_key):
        """API: Trigger snapshot refresh for a specific server (Admin Mass Refresh)"""
        try:
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Server key diperlukan'}), 400

            tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
            result = SnapshotManager.trigger_refresh(server_key, tanggal)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def trigger_delta_refresh_target(server_key):
        """API: Quick update (delta) for a specific server"""
        try:
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Server key diperlukan'}), 400

            tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
            result = SnapshotManager.trigger_delta_refresh(server_key, tanggal)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def trigger_weekly_refresh_target(server_key):
        """API: Weekly update for a specific server"""
        try:
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Server key diperlukan'}), 400

            tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
            result = SnapshotManager.trigger_weekly_refresh(server_key, tanggal)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def trigger_yearly_refresh_target(server_key):
        """API: Yearly update for a specific server"""
        try:
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Server key diperlukan'}), 400

            tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))
            result = SnapshotManager.trigger_yearly_refresh(server_key, tanggal)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def snapshot_status_target(server_key):
        """API: Get snapshot status for a specific server"""
        if not server_key:
            return jsonify({'state': 'empty', 'has_snapshot': False})

        status = SnapshotManager.get_status(server_key)
        return jsonify(status)

    @staticmethod
    def global_snapshot_status():
        """API: Get snapshot status for ALL servers (for global broadcasting)"""
        servers = db_manager.get_available_servers()
        statuses = {}
        for s in servers:
            sk = s['key']
            statuses[sk] = SnapshotManager.get_status(sk)
        return jsonify({'status': 'success', 'data': statuses})

    # ──────────── Data APIs ────────────

    @staticmethod
    def fetch_divisi_list():
        """API: Get distinct divisi names (lightweight — no full data load)"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            divisi_list = SnapshotManager.get_divisi_list(server_key)
            return jsonify({'status': 'success', 'divisi_list': divisi_list})

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_monitoring_data():
        """API: Search stok data from local snapshot (instant)"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            search_kode = request.args.get('search_kode')
            search_nama = request.args.get('search_nama')
            divisi = request.args.get('divisi')
            kategori = request.args.get('kategori')
            merk = request.args.get('merk')
            sort_by = request.args.get('sort_by')
            sort_order = request.args.get('sort_order', 'asc')

            # Pagination params
            limit = request.args.get('limit', type=int)
            offset = request.args.get('offset', type=int)

            result = SnapshotManager.search(
                server_key, search_kode, search_nama, divisi,
                kategori=kategori, merk=merk,
                limit=limit, offset=offset,
                sort_by=sort_by, sort_order=sort_order
            )
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_opname_data():
        """API: Search opname data from local snapshot"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            search_kode = request.args.get('search_kode')
            search_nama = request.args.get('search_nama')
            divisi = request.args.get('divisi')
            status = request.args.get('status')
            sort_by = request.args.get('sort_by')
            sort_order = request.args.get('sort_order', 'desc')

            result = SnapshotManager.search_opname(
                server_key, search_kode, search_nama, divisi,
                status=status, sort_by=sort_by, sort_order=sort_order
            )
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_item_stock_detail(kd_barang):
        """API: Get stock breakdown per divisi + satuan for one item"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            result = SnapshotManager.get_item_stock_detail(server_key, kd_barang)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_low_stock_alert():
        """API: Get low stock items from snapshot"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            min_stok = request.args.get('min_stok', 10, type=int)
            search_kode = request.args.get('search_kode')
            search_nama = request.args.get('search_nama')

            result = SnapshotManager.search(server_key, search_kode, search_nama)

            if result['status'] != 'success':
                return jsonify(result)

            low_stock = [
                row for row in result['data']
                if 0 < (row.get('Stok Akhir', 0) or 0) < min_stok
            ]

            return jsonify({
                'status': 'success',
                'data': low_stock,
                'row_count': len(low_stock),
                'threshold': min_stok
            })

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_barang_histori():
        """API: Get transaction history for one item in one division"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            kd_barang = request.args.get('kd_barang')
            kd_divisi = request.args.get('kd_divisi', '')
            start_date = request.args.get('start_date', '')
            end_date = request.args.get('end_date', '')

            if not kd_barang:
                return jsonify({'status': 'error', 'message': 'Kode barang harus diisi'}), 400

            result = SnapshotManager.get_barang_histori(server_key, kd_barang, kd_divisi, start_date, end_date)
            return jsonify(result)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_barang_tanpa_transaksi():
        """API: Get list of items with initial stock but no transactions"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            stok_filter = request.args.get('stok_filter', 'all')
            result = SnapshotManager.get_barang_tanpa_transaksi(server_key, stok_filter)
            return jsonify(result)
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_barang_dengan_transaksi():
        """API: Get list of items that have specific transactions in a year range"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            jenis_transaksi = request.args.get('jenis_transaksi', 'Semua')
            start_year = request.args.get('start_year')
            end_year = request.args.get('end_year')

            result = SnapshotManager.get_barang_dengan_transaksi(
                server_key, jenis_transaksi, start_year, end_year
            )
            return jsonify(result)
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_semua_barang_stok_awal():
        """API: Get list of ALL items with their initial stock"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            stok_filter = request.args.get('stok_filter', 'all')
            result = SnapshotManager.get_semua_barang_stok_awal(server_key, stok_filter)
            return jsonify(result)
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def export_xlsx():
        """API: Export search result as XLSX file"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            search_kode = request.args.get('search_kode')
            search_nama = request.args.get('search_nama')
            divisi = request.args.get('divisi')
            kategori = request.args.get('kategori')
            merk = request.args.get('merk')
            sort_by = request.args.get('sort_by')
            sort_order = request.args.get('sort_order', 'asc')

            result = SnapshotManager.search(
                server_key=server_key, 
                search_kode=search_kode, 
                search_nama=search_nama, 
                divisi=divisi,
                kategori=kategori,
                merk=merk,
                sort_by=sort_by,
                sort_order=sort_order
            )
            
            if result['status'] != 'success':
                return jsonify(result), 400

            data = result['data']
            
            from app.Services.StokService import StokService
            headers = [
                'Kode Divisi', 'Divisi', 'Kode Barang', 'Barang', 'Kategori', 
                'Merk', 'Model', 'Warna', 'Ukuran', 'Stok Akhir', 
                'Harga Average', 'Harga Jual', 'Nominal', 'Harga Beli Akhir'
            ]
            def mapping_fn(row):
                return [
                    row.get('Kode Divisi'), row.get('Divisi'), row.get('Kode Barang'), 
                    row.get('Barang'), row.get('Kategori'), row.get('Merk'), 
                    row.get('Model'), row.get('Warna'), row.get('Ukuran'), 
                    row.get('Stok Akhir'), row.get('Harga Average'), 
                    row.get('Harga Jual'), row.get('Nominal'), row.get('Harga Beli Akhir')
                ]
            output = StokService.generate_excel_from_data("Stok Monitoring", headers, data, mapping_fn)

            filename = f"stok_monitoring_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(
                output, 
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True, 
                download_name=filename
            )

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def export_histori_xlsx():
        """API: Export transaction history as XLSX file"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            kd_barang = request.args.get('kd_barang')
            kd_divisi = request.args.get('kd_divisi', '')
            start_date = request.args.get('start_date', '')
            end_date = request.args.get('end_date', '')

            if not kd_barang:
                return jsonify({'status': 'error', 'message': 'Kode barang harus diisi'}), 400

            result = SnapshotManager.get_barang_histori(server_key, kd_barang, kd_divisi, start_date, end_date)
            if result['status'] != 'success':
                return jsonify(result), 400

            data = result['data']
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Histori Barang"

            # Headers
            headers = [
                'Kd_Divisi', 'Divisi', 'K.Nota', 'Tanggal', 'Transaksi', 
                'No. Transaksi', 'Kd_Barang', 'Barang', 'Debet', 'Kredit', 
                'Kd_Satuan', 'Satuan', 'Harga', 'Saldo', 'Konversi'
            ]
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            use_konversi = request.args.get('konversi') == '1'

            # Data rows with running balance
            saldo = 0
            for row in data:
                conv = float(row.get('Konversi', 1) or 1) if use_konversi else 1.0
                debet = float(row.get('Debet', 0) or 0) * conv
                kredit = float(row.get('Kredit', 0) or 0) * conv
                saldo += (debet - kredit)
                
                satuan = row.get('satuan')
                if use_konversi and row.get('Konversi', 1) != 1.0:
                    satuan = f"{satuan} (Conv)"

                ws.append([
                    row.get('Kd_Divisi'), 
                    row.get('Divisi'), 
                    row.get('K.Nota'),
                    row.get('tanggal'), 
                    row.get('Transaksi'), 
                    row.get('no_transaksi'), 
                    row.get('kd_barang'), 
                    row.get('barang'), 
                    debet, 
                    kredit, 
                    row.get('kd_satuan'), 
                    satuan, 
                    row.get('harga'),
                    saldo,
                    row.get('Konversi', 1)
                ])

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Save to buffer
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"histori_{kd_barang}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(
                output, 
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True, 
                download_name=filename
            )

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def export_barang_tanpa_transaksi_xlsx():
        """API: Export list of items with initial stock but no transactions to XLSX"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            stok_filter = request.args.get('stok_filter', 'all')
            result = SnapshotManager.get_barang_tanpa_transaksi(server_key, stok_filter)
            if result['status'] != 'success':
                return jsonify(result), 400

            data = result['data']
            
            from app.Services.StokService import StokService
            headers = ['Kode Divisi', 'Kode Barang', 'Nama Barang', 'Stok Awal']
            def mapping_fn(row):
                return [
                    row.get('kd_divisi'), 
                    row.get('kd_barang'), 
                    row.get('nama_barang'),
                    row.get('stok_awal')
                ]
            output = StokService.generate_excel_from_data("Barang Tanpa Transaksi", headers, data, mapping_fn)

            filename = f"barang_tanpa_transaksi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(
                output, 
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True, 
                download_name=filename
            )

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def export_barang_dengan_transaksi_xlsx():
        """API: Export list of items with specific transactions to XLSX"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            jenis_transaksi = request.args.get('jenis_transaksi', 'Semua')
            start_year = request.args.get('start_year')
            end_year = request.args.get('end_year')

            result = SnapshotManager.get_barang_dengan_transaksi(
                server_key, jenis_transaksi, start_year, end_year
            )
            
            if result['status'] != 'success':
                return jsonify(result), 400

            data = result['data']
            
            from app.Services.StokService import StokService
            headers = ['Kode Divisi', 'Kode Barang', 'Nama Barang']
            def mapping_fn(row):
                return [
                    row.get('kd_divisi'), 
                    row.get('kd_barang'), 
                    row.get('nama_barang')
                ]
            output = StokService.generate_excel_from_data("Barang Dengan Transaksi", headers, data, mapping_fn)

            filename = f"barang_dengan_transaksi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(
                output, 
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True, 
                download_name=filename
            )

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def export_bulk_transaksi_xlsx():
        """API: Export bulk detail transactions row-by-row to XLSX"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            jenis_transaksi = request.args.get('jenis_transaksi', 'Semua')
            start_year = request.args.get('start_year')
            end_year = request.args.get('end_year')

            result = SnapshotManager.get_bulk_transaksi_detail(
                server_key, jenis_transaksi, start_year, end_year
            )
            
            if result['status'] != 'success':
                return jsonify(result), 400

            data = result['data']
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Detail Transaksi"

            # Headers
            headers = ['Divisi', 'Kd Barang', 'Nama Barang', 'Tanggal', 'Jenis Transaksi', 'No Transaksi', 'Qty', 'Satuan', 'Harga']
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Data rows
            for row in data:
                # Format Tanggal
                tanggal = row.get('tanggal')
                if isinstance(tanggal, str):
                    try:
                        tanggal_obj = datetime.strptime(tanggal, '%Y-%m-%d %H:%M:%S')
                        tanggal_str = tanggal_obj.strftime('%d/%m/%Y %H:%M')
                    except:
                        tanggal_str = tanggal
                elif tanggal:
                    tanggal_str = tanggal.strftime('%d/%m/%Y %H:%M')
                else:
                    tanggal_str = '-'

                ws.append([
                    row.get('nama_divisi') or row.get('kd_divisi'),
                    row.get('kd_barang'),
                    row.get('nama_barang'),
                    tanggal_str,
                    row.get('jenis_transaksi'),
                    row.get('no_transaksi'),
                    row.get('qty'),
                    row.get('nama_satuan') or row.get('kd_satuan'),
                    row.get('harga')
                ])

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Save to buffer
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"detail_{jenis_transaksi.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(
                output, 
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True, 
                download_name=filename
            )

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def export_semua_barang_stok_awal_xlsx():
        """API: Export list of ALL items with their initial stock to XLSX"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            stok_filter = request.args.get('stok_filter', 'all')
            result = SnapshotManager.get_semua_barang_stok_awal(server_key, stok_filter)
            if result['status'] != 'success':
                return jsonify(result), 400

            data = result['data']
            
            from app.Services.StokService import StokService
            headers = ['Kode Divisi', 'Kode Barang', 'Nama Barang', 'Stok Awal']
            def mapping_fn(row):
                return [
                    row.get('kd_divisi'), 
                    row.get('kd_barang'), 
                    row.get('nama_barang'),
                    row.get('stok_awal')
                ]
            output = StokService.generate_excel_from_data("Stok Awal Semua Barang", headers, data, mapping_fn)

            filename = f"semua_barang_stok_awal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(
                output, 
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True, 
                download_name=filename
            )

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    # ──────────── Update Barang APIs ────────────

    @staticmethod
    def fetch_barang_data():
        """API: Fetch data barang + harga + status dari MSSQL (live query)"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            q = request.args.get('q', '').strip()
            kategori = request.args.get('kategori', '').strip()
            merk = request.args.get('merk', '').strip()
            margin = request.args.get('margin', '').strip()
            status_filter = request.args.get('status', '').strip()
            last_kd = request.args.get('last_kd', '').strip()
            limit = int(request.args.get('limit', 20))
            q_type = request.args.get('q_type', 'kode').strip()

            sql = """
                SELECT TOP (?) 
                    b.kd_barang,
                    b.barang,
                    b.kategori,
                    b.merk,
                    s.harga_jual,
                    s.margin,
                    s.kd_satuan,
                    s.jumlah,
                    sat.satuan as nama_satuan,
                    mb.status as status_barang,
                    md.status as status_divisi,
                    s.status as status_satuan
                FROM dbo.v_m_barang b
                JOIN dbo.m_barang mb ON b.kd_barang = mb.kd_barang
                JOIN dbo.m_barang_satuan s ON b.kd_barang = s.kd_barang
                LEFT JOIN dbo.m_barang_divisi md ON b.kd_barang = md.kd_barang
                LEFT JOIN dbo.v_m_barang_satuan sat ON s.kd_satuan = sat.kd_satuan AND sat.kd_barang = b.kd_barang
                WHERE 1=1
            """
            params = [limit]

            if q:
                if q_type == 'nama':
                    sql += " AND b.barang LIKE ?"
                else: # Default to kode
                    sql += " AND b.kd_barang LIKE ?"
                params.append(f'%{q}%')
            if kategori:
                sql += " AND b.kategori LIKE ?"
                params.append(f'%{kategori}%')
            if merk:
                sql += " AND b.merk LIKE ?"
                params.append(f'%{merk}%')
            if margin:
                if margin == 'high':
                    sql += " AND s.margin >= 30"
                elif margin == 'medium':
                    sql += " AND s.margin >= 15 AND s.margin < 30"
                elif margin == 'low':
                    sql += " AND s.margin < 15"
            if status_filter:
                if status_filter == 'nonaktif':
                    sql += " AND (mb.status = 0 OR md.status = 0 OR s.status = 0)"
                elif status_filter == 'aktif':
                    sql += " AND (mb.status = 1 OR md.status = 1 OR s.status = 1)"
                    
            if last_kd:
                sql += " AND b.kd_barang > ?"
                params.append(last_kd)

            sql += " ORDER BY b.kd_barang"

            rows = db_manager.execute_query(server_key, sql, tuple(params))
            return jsonify(rows)

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def update_barang():
        """API POST: Update harga jual (+ margin) dan 3 field status barang"""
        try:
            server_key = session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            data = request.get_json() or {}
            kd_barang = data.get('kd_barang')
            kd_satuan = data.get('kd_satuan')
            harga_jual_baru = data.get('harga_jual')
            
            # Statuses
            status_barang = data.get('status_barang')
            status_divisi = data.get('status_divisi')
            status_satuan = data.get('status_satuan')

            if not all([kd_barang, kd_satuan, harga_jual_baru is not None, status_barang is not None, status_divisi is not None, status_satuan is not None]):
                return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400

            try:
                harga_jual_baru = float(harga_jual_baru)
                status_barang = int(status_barang)
                status_divisi = int(status_divisi)
                status_satuan = int(status_satuan)
            except ValueError:
                return jsonify({'status': 'error', 'message': 'Format data tidak valid'}), 400

            # Get server type
            servers = db_manager.get_available_servers()
            server = next((s for s in servers if s['key'] == server_key), None)
            server_type = server.get('type', 'grosir') if server else 'grosir'

            # Get current price and margin
            sql_check = """
                SELECT harga_jual, margin
                FROM dbo.m_barang_satuan
                WHERE kd_barang = ? AND kd_satuan = ?
            """
            result = db_manager.execute_query(server_key, sql_check, (kd_barang, kd_satuan))

            if not result:
                return jsonify({'status': 'error', 'message': 'Produk tidak ditemukan'}), 404

            row = result[0]
            harga_jual_lama = float(row['harga_jual'])
            margin_lama = float(row['margin'] or 0)

            if server_type == 'eceran':
                # Eceran: Calculate margin from base price
                if margin_lama > 0:
                    harga_beli = harga_jual_lama / (1 + margin_lama / 100)
                else:
                    harga_beli = harga_jual_lama

                if harga_jual_baru < harga_beli:
                    return jsonify({
                        'status': 'error',
                        'message': f'Harga baru (Rp {harga_jual_baru:,.0f}) tidak boleh lebih rendah dari harga beli (Rp {harga_beli:,.0f})'
                    }), 400

                margin_baru = ((harga_jual_baru - harga_beli) / harga_beli) * 100 if harga_beli > 0 else 0
            else:
                # Grosir: margin selalu 0
                harga_beli = harga_jual_lama
                margin_baru = 0

            # 1. Update m_barang
            db_manager.execute_update(server_key, "UPDATE dbo.m_barang SET status = ? WHERE kd_barang = ?", (status_barang, kd_barang))
            
            # 2. Update m_barang_divisi
            db_manager.execute_update(server_key, "UPDATE dbo.m_barang_divisi SET status = ? WHERE kd_barang = ?", (status_divisi, kd_barang))

            # 3. Update m_barang_satuan (Harga + Status)
            sql_update_satuan = """
                UPDATE dbo.m_barang_satuan
                SET harga_jual = ?, margin = ?, status = ?
                WHERE kd_barang = ? AND kd_satuan = ?
            """
            db_manager.execute_update(server_key, sql_update_satuan, (
                harga_jual_baru, margin_baru, status_satuan, kd_barang, kd_satuan
            ))

            return jsonify({
                'status': 'ok',
                'message': 'Barang berhasil diupdate',
                'data': {
                    'kd_barang': kd_barang,
                    'kd_satuan': kd_satuan,
                    'harga_jual': harga_jual_baru,
                    'margin': margin_baru,
                    'harga_beli': harga_beli,
                    'profit': harga_jual_baru - harga_beli,
                    'status_barang': status_barang,
                    'status_divisi': status_divisi,
                    'status_satuan': status_satuan
                }
            })

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    # ──────────── Compare & Sync Harga Gudang APIs ────────────

    @staticmethod
    def compare_harga_page():
        """Render halaman komparasi harga Server Gudang vs Server Grosir"""
        servers = db_manager.get_available_servers()
        
        gudang_servers = [s for s in servers if s.get('type') == 'gudang']
        grosir_servers = [s for s in servers if s.get('type') == 'grosir']
        
        server_key = session.get('selected_server')
        return render_template('compare_harga.html', 
                               gudang_servers=gudang_servers, 
                               grosir_servers=grosir_servers, 
                               selected_server=server_key)

    @staticmethod
    def fetch_compare_harga():
        """API: Ambil data mismatch harga Source vs Target Server"""
        try:
            source_server = request.args.get('source_server', 'GUDANG')
            target_server = request.args.get('target_server')

            if not target_server or source_server == target_server:
                return jsonify({'status': 'error', 'message': 'Pilih server target yang berbeda dari source'}), 400

            # Query dasar untuk ngambil harga
            sql = """
                SELECT 
                    b.kd_barang, 
                    b.barang, 
                    s.kd_satuan, 
                    sat.satuan as nama_satuan,
                    s.harga_jual,
                    mb.status as status_barang
                FROM dbo.v_m_barang b
                JOIN dbo.m_barang mb ON b.kd_barang = mb.kd_barang
                JOIN dbo.m_barang_satuan s ON b.kd_barang = s.kd_barang
                LEFT JOIN dbo.v_m_barang_satuan sat ON s.kd_satuan = sat.kd_satuan AND sat.kd_barang = b.kd_barang
            """

            # Fetch Source
            source_rows = db_manager.execute_query(source_server, sql)
            if not source_rows:
                return jsonify({'status': 'success', 'data': []})
                
            # Build dict for Source: key = (kd_barang, kd_satuan)
            source_dict = {
                (row['kd_barang'], row['kd_satuan']): row 
                for row in source_rows
            }

            # Fetch Target Server
            target_rows = db_manager.execute_query(target_server, sql)
            target_dict = {
                (row['kd_barang'], row['kd_satuan']): row 
                for row in target_rows
            }

            mismatches = []
            for key, g_row in source_dict.items():
                if key in target_dict:
                    t_row = target_dict[key]
                    h_source = float(g_row['harga_jual'] or 0)
                    h_target = float(t_row['harga_jual'] or 0)
                    
                    if h_source != h_target:
                        mismatches.append({
                            'kd_barang': g_row['kd_barang'],
                            'nama_barang': g_row['barang'],
                            'kd_satuan': g_row['kd_satuan'],
                            'nama_satuan': g_row['nama_satuan'],
                            'harga_gudang': h_source,
                            'harga_target': h_target,
                            'selisih': abs(h_source - h_target),
                            'status_barang': g_row['status_barang']
                        })

            return jsonify({'status': 'success', 'data': mismatches})
            
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def sync_harga_gudang():
        """API POST: Sync harga target server dari list items ke harga source"""
        import os, json
        from datetime import datetime
        try:
            data = request.get_json() or {}
            target_server = data.get('target_server')
            items = data.get('items', [])
            
            if not target_server:
                return jsonify({'status': 'error', 'message': 'Target server tidak valid'}), 400
            
            if not items:
                return jsonify({'status': 'error', 'message': 'Tidak ada item yang dipilih'}), 400

            success_count = 0
            for item in items:
                kd_barang = item.get('kd_barang')
                kd_satuan = item.get('kd_satuan')
                harga_gudang = float(item.get('harga_gudang', 0))
                
                sql_update = """
                    UPDATE dbo.m_barang_satuan
                    SET harga_jual = ?
                    WHERE kd_barang = ? AND kd_satuan = ?
                """
                db_manager.execute_update(target_server, sql_update, (harga_gudang, kd_barang, kd_satuan))
                success_count += 1
                
            # Log history
            if success_count > 0:
                history_file = os.path.join('database', 'sync_history.json')
                history_data = []
                if os.path.exists(history_file):
                    try:
                        with open(history_file, 'r', encoding='utf-8') as f:
                            history_data = json.load(f)
                    except:
                        pass
                
                # Determine source from items (usually same for all)
                source_server = data.get('source_server', 'GUDANG')
                
                history_data.insert(0, {
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'user': session.get('username', 'Admin'),
                    'source': source_server,
                    'target': target_server,
                    'count': success_count,
                    'items': [f"{i.get('kd_barang')} ({i.get('nama_barang')})" for i in items][:10]
                })
                
                history_data = history_data[:50] # Keep last 50
                
                with open(history_file, 'w', encoding='utf-8') as f:
                    json.dump(history_data, f, indent=2)
                
            return jsonify({'status': 'success', 'message': f'Berhasil sync {success_count} item'})

        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def fetch_sync_history():
        """API GET: Fetch price sync history"""
        import os, json
        try:
            history_file = os.path.join('database', 'sync_history.json')
            history_data = []
            if os.path.exists(history_file):
                with open(history_file, 'r', encoding='utf-8') as f:
                    history_data = json.load(f)
            return jsonify({'status': 'success', 'data': history_data})
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    # ──────────── DuckDB Sync APIs ────────────

    @staticmethod
    def sync_duckdb_page():
        """Render halaman khusus untuk sinkronisasi DuckDB"""
        server_key = session.get('selected_server')
        if not server_key:
            flash('Pilih server terlebih dahulu', 'error')
            return redirect(url_for('web.dashboard_page'))
            
        servers = db_manager.get_available_servers()
        if session.get('role') != 'super_admin':
            allowed_servers = session.get('servers', [])
            servers = [s for s in servers if s['key'] in allowed_servers]
            
        return render_template('sync_duckdb.html', 
                             servers=servers,
                             selected_server=server_key,
                             role=session.get('role'))

    @staticmethod
    def trigger_duckdb_sync():
        """API: Trigger background sync untuk DuckDB"""
        try:
            data = request.get_json() or {}
            server_key = data.get('server_key') or session.get('selected_server')
            if not server_key:
                return jsonify({'status': 'error', 'message': 'Pilih server terlebih dahulu'}), 400

            days_back = data.get('days_back', 30)
            
            from app.Services.Snapshot.SalesDuckDBRunner import SalesDuckDBRunner
            result = SalesDuckDBRunner.sync_sales_data(server_key, days_back=days_back)
            return jsonify(result)
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def check_duckdb_status():
        """API: Cek status sync DuckDB"""
        server_key = session.get('selected_server')
        if not server_key:
            return jsonify({'state': 'empty'})

        from app.Services.Snapshot.SalesDuckDBRunner import SalesDuckDBRunner
        status = SalesDuckDBRunner.get_status(server_key)
        return jsonify(status)

    @staticmethod
    def mass_sync_duckdb_page():
        return render_template('mass_sync_duckdb.html')

    @staticmethod
    def trigger_mass_duckdb_sync():
        try:
            data = request.get_json() or {}
            days_back = int(data.get('days_back', 30))
            servers = db_manager.get_available_servers()
            from app.Services.Snapshot.SalesDuckDBRunner import SalesDuckDBRunner
            for srv in servers:
                if srv.get('type') == 'retail':
                    continue
                SalesDuckDBRunner.sync_sales_data(srv['key'], days_back=days_back)
            return jsonify({'status': 'success', 'message': 'Mass sync dimulai untuk semua server'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def check_mass_duckdb_status():
        try:
            from app.Services.Snapshot.SalesDuckDBRunner import SalesDuckDBRunner
            status = SalesDuckDBRunner.get_all_status()
            return jsonify({'status': 'success', 'data': status})
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def dashboard_analytics_trend():
        server_key = session.get('selected_server')
        tahun = request.args.get('tahun', datetime.now().year)
        if not server_key: return jsonify({'error': 'No server'}), 400
        from app.Services.DashboardAnalyticsService import DashboardAnalyticsService
        return jsonify(DashboardAnalyticsService.get_monthly_trend(server_key, int(tahun)))

    @staticmethod
    def dashboard_analytics_retention():
        server_key = session.get('selected_server')
        tahun = request.args.get('tahun', datetime.now().year)
        if not server_key: return jsonify({'error': 'No server'}), 400
        from app.Services.DashboardAnalyticsService import DashboardAnalyticsService
        return jsonify(DashboardAnalyticsService.get_customer_retention(server_key, int(tahun)))

    @staticmethod
    def dashboard_analytics_heatmap():
        server_key = session.get('selected_server')
        tahun = request.args.get('tahun', datetime.now().year)
        if not server_key: return jsonify({'error': 'No server'}), 400
        from app.Services.DashboardAnalyticsService import DashboardAnalyticsService
        return jsonify(DashboardAnalyticsService.get_traffic_heatmap(server_key, int(tahun)))

    @staticmethod
    def dashboard_analytics_validate():
        server_key = session.get('selected_server')
        tahun = request.args.get('tahun', datetime.now().year)
        if not server_key: return jsonify({'error': 'No server'}), 400
        from app.Services.DataValidationService import DataValidationService
        return jsonify(DataValidationService.validate_accuracy(server_key, int(tahun)))

    @staticmethod
    def dashboard_analytics_radar():
        tahun = request.args.get('tahun', datetime.now().year)
        from app.Services.DashboardAnalyticsService import DashboardAnalyticsService
        return jsonify(DashboardAnalyticsService.get_cross_branch_omset(int(tahun)))

    @staticmethod
    def dashboard_analytics_basket():
        server_key = session.get('selected_server')
        tahun = request.args.get('tahun', datetime.now().year)
        if not server_key: return jsonify({'error': 'No server'}), 400
        from app.Services.DashboardAnalyticsService import DashboardAnalyticsService
        return jsonify(DashboardAnalyticsService.get_basket_composition(server_key, int(tahun)))

    # ──────────── Perhitungan Stok APIs ────────────

    @staticmethod
    def trigger_perhitungan_stok():
        """API: Trigger perhitungan stok di background"""
        try:
            data = request.get_json() or {}
            server_key = session.get('selected_server')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            use_stok_awal = data.get('use_stok_awal', False)

            if not server_key or not start_date or not end_date:
                return jsonify({'status': 'error', 'message': 'Missing parameters'}), 400

            from app.Services.Snapshot.SnapshotRunner import SnapshotRunner
            result = SnapshotRunner.trigger_perhitungan_stok(server_key, start_date, end_date, use_stok_awal)
            return jsonify(result)
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def trigger_perhitungan_stok_tanggal():
        """API: Trigger perhitungan stok berdasarkan 1 tanggal persis spt monitoring"""
        try:
            data = request.get_json() or {}
            server_key = session.get('selected_server')
            tanggal = data.get('tanggal')

            if not server_key or not tanggal:
                return jsonify({'status': 'error', 'message': 'Missing parameters'}), 400

            from app.Services.Snapshot.SnapshotRunner import SnapshotRunner
            result = SnapshotRunner.trigger_perhitungan_stok(server_key, start_date='1900-01-01', end_date=tanggal, use_stok_awal=True)
            return jsonify(result)
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @staticmethod
    def status_perhitungan_stok():
        """API: Cek status perhitungan stok"""
        server_key = session.get('selected_server')
        if not server_key:
            return jsonify({'state': 'empty'})

        from app.Services.Snapshot.SnapshotState import SnapshotState
        status = SnapshotState._perhitungan_status.get(server_key)
        if not status:
            return jsonify({'state': 'empty'})
        return jsonify(status)

    @staticmethod
    def fetch_perhitungan_stok():
        """API: Fetch data perhitungan stok"""
        server_key = session.get('selected_server')
        if not server_key:
            return jsonify({'status': 'error', 'message': 'Pilih server dulu'}), 400

        search_kode = request.args.get('search_kode')
        search_nama = request.args.get('search_nama')
        divisi = request.args.get('divisi')
        kategori = request.args.get('kategori')
        merk = request.args.get('merk')
        
        limit = request.args.get('limit', type=int)
        offset = request.args.get('offset', type=int)
            
        sort_by = request.args.get('sort_by', 'nominal')
        sort_order = request.args.get('sort_order', 'desc')

        from app.Services.Snapshot.SnapshotQuery import SnapshotQuery
        result = SnapshotQuery.search_perhitungan(
            server_key=server_key,
            search_kode=search_kode,
            search_nama=search_nama,
            divisi=divisi,
            kategori=kategori,
            merk=merk,
            limit=limit,
            offset=offset,
            sort_by=sort_by,
            sort_order=sort_order
        )

        return jsonify(result)

    @staticmethod
    def export_perhitungan_stok_xlsx():
        """API: Export perhitungan stok ke XLSX"""
        server_key = session.get('selected_server')
        if not server_key:
            return "Pilih server dulu", 400

        search_kode = request.args.get('search_kode')
        search_nama = request.args.get('search_nama')
        divisi = request.args.get('divisi')
        kategori = request.args.get('kategori')
        merk = request.args.get('merk')

        from app.Services.Snapshot.SnapshotQuery import SnapshotQuery
        result = SnapshotQuery.search_perhitungan(
            server_key=server_key,
            search_kode=search_kode,
            search_nama=search_nama,
            divisi=divisi,
            kategori=kategori,
            merk=merk,
            limit=1000000,
            offset=0
        )

        if result.get('status') != 'success':
            return result.get('message', 'Failed to fetch data'), 400

        data = result['data']
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Perhitungan Stok"

        headers = ['Kode Divisi', 'Divisi', 'Kode Barang', 'Barang', 'Kategori', 'Merk', 'Stok Akhir', 'Harga Avg', 'Harga Jual', 'Nominal', 'Harga Beli Akhir']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row in data:
            ws.append([
                row.get('Kode Divisi'), row.get('Divisi'), row.get('Kode Barang'),
                row.get('Barang'), row.get('Kategori'), row.get('Merk'),
                row.get('Stok Akhir'), row.get('Harga Avg'), row.get('Harga Jual'),
                row.get('Nominal'), row.get('Harga Beli Akhir')
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"Perhitungan_Stok_{server_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @staticmethod
    def dashboard_analytics_stock_predict():
        server_key = session.get('selected_server')
        if not server_key: return jsonify({'error': 'No server'}), 400
        from app.Services.DashboardAnalyticsService import DashboardAnalyticsService
        return jsonify(DashboardAnalyticsService.get_stock_prediction(server_key))

